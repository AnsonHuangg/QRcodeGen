'''For QR code generating'''
import os

import qrcode
from qrcode.image.styledpil import StyledPilImage

from openpyxl import load_workbook

#===================================================================================================
#                                           Forder check
#===================================================================================================
####################################################################################################
# Path want to check
####################################################################################################
directoryNow = os.path.dirname(os.path.abspath(__file__))

folderPathOutput = directoryNow + "\\Output"
folderPathInput = directoryNow + "\\Input"

####################################################################################################
# Check whether the folder is exist. If no, auto generate
####################################################################################################
if not os.path.isdir(folderPathOutput):
    os.makedirs(folderPathOutput)

#===================================================================================================
#                                           Excel analyze
#===================================================================================================
####################################################################################################
# Read Excel file
####################################################################################################
wb = load_workbook('data.xlsx')

####################################################################################################
# Get Sheet by name
####################################################################################################
sheet = wb['Sheet1']

#===================================================================================================
#                                           QR code generate
#===================================================================================================
for i in range(2, sheet.max_row + 1):
    ################################################################################################
    # Get excel data
    ################################################################################################
    numData  = sheet.cell(row=i, column=1).value
    moteData = sheet.cell(row=i, column=2).value
    linkData = sheet.cell(row=i, column=3).value
    inNameData = sheet.cell(row=i, column=4).value
    outNameData = sheet.cell(row=i, column=5).value

    ################################################################################################
    # input/output file name assign
    ################################################################################################
    inputFileName = folderPathInput + "\\" + inNameData
    outputFileName = folderPathOutput + "\\" + str(outNameData) + ".jpg"

    ################################################################################################
    # QR Code gen
    #
    # Some parameter to config
    # 1. box_size
    #       Set the nunber of pixel for each QR code box
    # 2. border
    #       Set the width of QR code border
    # 3. version
    #        Set the QR code version
    # 4. Error correction
    #        ERROR_CORRECT_L: Can correct the mistake less than 7%
    #        ERROR_CORRECT_M: Can correct the mistake less than 15% (default)
    #        ERROR_CORRECT_Q: Can correct the mistake less than 25%
    #        ERROR_CORRECT_H: Can correct the mistake less than 30%
    ################################################################################################
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H)
    qr.add_data(linkData)
    img = qr.make_image(image_factory=StyledPilImage, embeded_image_path=inputFileName)

    ################################################################################################
    # Check output file is exist or not. If yes, remove it
    ################################################################################################
    if os.path.isfile(outputFileName):
        os.remove(outputFileName)

    ################################################################################################
    # Create file
    ################################################################################################
    img.save(outputFileName)

#===================================================================================================
#                                            EOF
#===================================================================================================
